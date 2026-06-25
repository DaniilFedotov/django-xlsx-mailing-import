from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from openpyxl import load_workbook

from mailings.constants import REQUIRED_COLUMNS, USER_ID_INVALID, USER_ID_NOT_POSITIVE
from mailings.exceptions import MissingColumnsError
from mailings.models import MailingRecord
from mailings.services import send_email
from mailings.types import ImportStats, MailingRow


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_row(row_values: tuple[object, ...], column_map: dict[str, int]) -> MailingRow:
    data: dict[str, str] = {}
    for column_name, index in column_map.items():
        value = row_values[index] if index < len(row_values) else None
        data[column_name] = "" if value is None else str(value).strip()
    return data  # type: ignore[return-value]


def _validate_row(data: MailingRow) -> None:
    for field_name in REQUIRED_COLUMNS:
        if not data.get(field_name):
            raise ValidationError(f"Missing required field: {field_name}")

    validate_email(data["email"])

    try:
        user_id = int(data["user_id"])
    except (TypeError, ValueError) as exc:
        raise ValidationError(USER_ID_INVALID) from exc

    if user_id <= 0:
        raise ValidationError(USER_ID_NOT_POSITIVE)


def _is_empty_row(row_values: tuple[object, ...] | None) -> bool:
    if row_values is None:
        return True
    return all(value is None or str(value).strip() == "" for value in row_values)


def import_mailings_from_xlsx(file_path: str) -> ImportStats:
    """Import mailing rows from an XLSX file and send emails for new records."""
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header_row = next(rows, None)

    if not header_row:
        workbook.close()
        return ImportStats()

    headers = [_normalize_header(value) for value in header_row]
    column_map = {name: index for index, name in enumerate(headers) if name}

    missing_columns = [column for column in REQUIRED_COLUMNS if column not in column_map]
    if missing_columns:
        workbook.close()
        raise MissingColumnsError(f"Missing required columns: {', '.join(missing_columns)}")

    stats = ImportStats()

    for row_values in rows:
        if _is_empty_row(row_values):
            continue

        stats.processed += 1
        row_data = _parse_row(row_values, column_map)

        try:
            _validate_row(row_data)
        except ValidationError:
            stats.errors += 1
            continue

        external_id = row_data["external_id"]
        if MailingRecord.objects.filter(external_id=external_id).exists():
            stats.skipped += 1
            continue

        send_email(
            email=row_data["email"],
            subject=row_data["subject"],
            message=row_data["message"],
        )
        MailingRecord.objects.create(
            external_id=external_id,
            user_id=int(row_data["user_id"]),
            email=row_data["email"],
            subject=row_data["subject"],
            message=row_data["message"],
        )
        stats.created += 1

    workbook.close()
    return stats
